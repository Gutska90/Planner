"""
Pipelines para procesar y exportar items
"""
import json
import os
from datetime import datetime
from itemadapter import ItemAdapter

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class JsonPipeline:
    """Pipeline para exportar a JSON"""
    
    def __init__(self):
        self.items = []
        self.filename = None

    def open_spider(self, spider):
        """Abrir spider"""
        # Determinar nombre del archivo basado en el nombre del spider
        if spider.name == 'carnes-y-pescados':
            self.filename = 'carnes-y-pescados_products.json'
        elif spider.name == 'destilados':
            self.filename = 'destilados_products.json'
        elif spider.name == 'snacks-y-picoteo':
            self.filename = 'snacks-y-picoteo_products.json'
        else:
            self.filename = f'{spider.name}_products.json'
        
        self.items = []
        spider.logger.info(f'Pipeline JSON inicializado: {self.filename}')

    def close_spider(self, spider):
        """Cerrar spider y guardar JSON"""
        if self.items:
            # Filtrar items de debug/blocked
            valid_items = [item for item in self.items if not item.get('_blocked') and not item.get('_debug')]
            
            # Eliminar duplicados basándose en product_url (campo único)
            seen_urls = set()
            unique_items = []
            duplicates_count = 0
            
            for item in valid_items:
                product_url = item.get('product_url')
                if product_url and product_url not in seen_urls:
                    seen_urls.add(product_url)
                    unique_items.append(item)
                elif product_url:
                    duplicates_count += 1
            
            if duplicates_count > 0:
                spider.logger.info(f'🔄 Eliminados {duplicates_count} productos duplicados')
            
            output_file = os.path.join(os.getcwd(), self.filename)
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(unique_items, f, ensure_ascii=False, indent=2)
            
            spider.logger.info(f'✅ Datos guardados en {output_file}: {len(unique_items)} productos únicos')
        else:
            spider.logger.warning('No hay productos para guardar en JSON')

    def process_item(self, item, spider):
        """Procesar item"""
        adapter = ItemAdapter(item)
        item_dict = dict(adapter)
        self.items.append(item_dict)
        return item


class ExcelPipeline:
    """Pipeline para exportar a Excel"""
    
    def __init__(self):
        self.items = []
        self.filename = None

    def open_spider(self, spider):
        """Abrir spider"""
        if not OPENPYXL_AVAILABLE:
            spider.logger.warning('openpyxl no disponible, desactivando ExcelPipeline')
            return
        
        # Determinar nombre del archivo basado en el nombre del spider
        if spider.name == 'carnes-y-pescados':
            self.filename = 'carnes-y-pescados_products.xlsx'
        elif spider.name == 'destilados':
            self.filename = 'destilados_products.xlsx'
        elif spider.name == 'snacks-y-picoteo':
            self.filename = 'snacks-y-picoteo_products.xlsx'
        else:
            self.filename = f'{spider.name}_products.xlsx'
        
        self.items = []
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Productos Lider"
        
        # Headers
        headers = ['URL Categoría', 'Nombre', 'Precio', 'Precio Descuento', 'URL Producto']
        for idx, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar ancho de columnas
        column_widths = [50, 60, 15, 15, 50]
        for idx, width in enumerate(column_widths, 1):
            self.ws.column_dimensions[get_column_letter(idx)].width = width
        
        spider.logger.info(f'Pipeline Excel inicializado: {self.filename}')

    def close_spider(self, spider):
        """Cerrar spider y guardar Excel"""
        if not OPENPYXL_AVAILABLE:
            return
        
        if self.items:
            # Filtrar items válidos
            valid_items = [item for item in self.items if not item.get('_blocked') and not item.get('_debug')]
            
            # Eliminar duplicados basándose en product_url
            seen_urls = set()
            unique_items = []
            duplicates_count = 0
            
            for item in valid_items:
                product_url = item.get('product_url')
                if product_url and product_url not in seen_urls:
                    seen_urls.add(product_url)
                    unique_items.append(item)
                elif product_url:
                    duplicates_count += 1
            
            if duplicates_count > 0:
                spider.logger.info(f'🔄 Eliminados {duplicates_count} productos duplicados del Excel')
            
            # Reconstruir el workbook con items únicos
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Productos Lider"
            
            # Headers
            headers = ['URL Categoría', 'Nombre', 'Precio', 'Precio Descuento', 'URL Producto']
            for idx, header in enumerate(headers, 1):
                cell = self.ws.cell(row=1, column=idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ajustar ancho de columnas
            column_widths = [50, 60, 15, 15, 50]
            for idx, width in enumerate(column_widths, 1):
                self.ws.column_dimensions[get_column_letter(idx)].width = width
            
            # Agregar items únicos
            for item in unique_items:
                row = len(unique_items) - unique_items.index(item) + 1  # +1 por el header
                self.ws.cell(row=row, column=1, value=item.get('category_url', ''))
                self.ws.cell(row=row, column=2, value=item.get('name', ''))
                self.ws.cell(row=row, column=3, value=item.get('price', ''))
                self.ws.cell(row=row, column=4, value=item.get('discount_price', ''))
                
                # URL como hipervínculo
                url_cell = self.ws.cell(row=row, column=5, value=item.get('product_url', ''))
                if item.get('product_url'):
                    url_cell.hyperlink = item.get('product_url')
                    url_cell.font = Font(underline="single", color="0563C1")
            
            output_file = os.path.join(os.getcwd(), self.filename)
            self.wb.save(output_file)
            
            spider.logger.info(f'✅ Datos guardados en {output_file}: {len(unique_items)} productos únicos')
        else:
            spider.logger.warning('No hay productos para guardar en Excel')

    def process_item(self, item, spider):
        """Procesar item"""
        if not OPENPYXL_AVAILABLE:
            return item
        
        adapter = ItemAdapter(item)
        
        # Saltar items de debug/blocked
        if adapter.get('_blocked') or adapter.get('_debug'):
            return item
        
        self.items.append(dict(adapter))
        
        # Agregar fila al Excel
        row = len(self.items) + 1  # +1 por el header
        
        self.ws.cell(row=row, column=1, value=adapter.get('category_url', ''))
        self.ws.cell(row=row, column=2, value=adapter.get('name', ''))
        self.ws.cell(row=row, column=3, value=adapter.get('price', ''))
        self.ws.cell(row=row, column=4, value=adapter.get('discount_price', ''))
        
        # URL como hipervínculo
        url_cell = self.ws.cell(row=row, column=5, value=adapter.get('product_url', ''))
        if adapter.get('product_url'):
            url_cell.hyperlink = adapter.get('product_url')
            url_cell.font = Font(underline="single", color="0563C1")
        
        return item

